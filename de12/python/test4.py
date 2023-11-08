import openpyxl

#新しいワークブックを作成
wb = openpyxl.Workbook()
#ワークブック内のアクティブなシートを取得
ws = wb.active
#ワークブックを名前をつけて保存
wb.save("家計簿リスト２.xlsx")

import openpyxl

wb = openpyxl.load_workbook('家計簿リスト２.xlsx')
ws = wb.active

#配列宣言
day = ["日付",input('日付')]
item = ["品目",input('品目')]
money =["金額",int(input('金額'))]

#配列ループ2行目
for i in range(0,len(day)):
    #A列にリストを書き込み
    ws.cell(i+1,1,value = day[i])

    #B列にリストを書き込み
    ws.cell(i+1,2,value = item[i])

    #c列にリストを書き込み
    ws.cell(i+1,3,value = money[i])
    
    i = i+1

select=input("書き足しますか？yes or no")
while select =="yes":
    import openpyxl
    wb = openpyxl.load_workbook('家計簿リスト２.xlsx')
    ws = wb.active


    #配列宣言
    day = [input('日付')]
    item = [input('品目')]
    money =[int(input('金額'))]

    #最大行
    maxRow = ws.max_row + 1

    #1列目を指定
    j = 1

    #行を逆ループ
    for i in reversed(range(1,maxRow)):

        #セルの値がNoneじゃなかったら、次の行から書き込み開始
        if ws.cell(row=i, column=j).value != None:

            #配列ループ3行目以降
            for Q in range(0,len(day)):

                #リストを書き込み
                ws.cell(i+1,j,value = day[Q])
                ws.cell(i+1,j+1,value = item[Q])
                ws.cell(i+1,j+2,value = money[Q])
                i = i + 1
        
        #保存        
            wb.save('家計簿リスト２.xlsx')
            select=input("書き足しますか？yes or no")