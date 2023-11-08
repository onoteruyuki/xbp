import openpyxl

wb = openpyxl.load_workbook('家計簿リスト２.xlsx')
ws = wb.active

#配列宣言
day = ["日付",input('日付')]
item = ["品目",input('品目')]
money =["金額",int(input('金額'))]

#配列ループ2行目
for i in range(0,len(day)):
    #A列にリストを書き込み
    ws.cell(i+1,1,value = day[i])

    #B列にリストを書き込み
    ws.cell(i+1,2,value = item[i])

    #c列にリストを書き込み
    ws.cell(i+1,3,value = money[i])
    
    i = i+1

#保存
wb.save('家計簿リスト２.xlsx')